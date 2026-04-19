from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from database import get_db
from auth import get_current_user
from models import WeeklyPlanCreate, TaskApprovalUpdate
from zoneinfo import ZoneInfo
from datetime import datetime, date as date_mod, timedelta

router = APIRouter()
bkk_tz = ZoneInfo('Asia/Bangkok')


def _now_bkk() -> str:
    return datetime.now(bkk_tz).strftime("%Y-%m-%d %H:%M:%S")


def _get_monday(date_str: str) -> str:
    """คืน YYYY-MM-DD ของวันจันทร์ในสัปดาห์ของ date_str"""
    d = date_mod.fromisoformat(date_str)
    monday = d - timedelta(days=d.weekday())
    return monday.isoformat()


def _get_week_dates(monday_str: str) -> list[str]:
    """คืน list 6 วัน (จ–ส) จาก monday_str"""
    monday = date_mod.fromisoformat(monday_str)
    return [(monday + timedelta(days=i)).isoformat() for i in range(6)]


def _get_subordinate_user_ids(db, current_user: dict) -> set | None:
    """
    คืน set ของ user_id ที่ admin/supervisor คนนี้ดูได้
    None หมายถึง super admin (ดูได้ทุกคน)
    """
    level = current_user.get("level", 0)
    role = current_user.get("role", "").lower()
    personal_id = current_user.get("user_id")
    is_super_admin = level == 9 or 'admin' in role
    if is_super_admin:
        return None
    allowed = {personal_id} if personal_id else set()
    if 1 <= level <= 3 and personal_id:
        for e in db.collection("evaluations") \
                    .where("evaluator_ids", "array_contains", personal_id) \
                    .stream():
            tid = e.to_dict().get("target_id")
            if tid:
                allowed.add(tid)
    return allowed


def _days_to_tasks(days: dict) -> list:
    """
    #? Backward compat: แปลง schema เก่า (days dict) → tasks list (schema ใหม่)
    แต่ละ task จะได้รับ active_days = [วันที่เดียวที่เคยอยู่]
    """
    seen_ids = {}
    result = []
    for date_str, task_list in days.items():
        for t in task_list:
            tid = t.get("id")
            if tid in seen_ids:
                # task ID เดิม → append วันใหม่เข้า active_days
                seen_ids[tid]["active_days"].append(date_str)
            else:
                task_copy = dict(t)
                task_copy["active_days"] = [date_str]
                seen_ids[tid] = task_copy
                result.append(task_copy)
    return result


# ─────────────────────────────────────────────────────────────────
# GET /api/plans  — ดูแผนงานของตัวเอง สำหรับสัปดาห์ที่ระบุ
# ─────────────────────────────────────────────────────────────────
@router.get("/")
def get_my_plan(
    week: str | None = None,
    current_user: dict = Depends(get_current_user)
):
    """
    คืนแผนงานของผู้ใช้ปัจจุบันสำหรับสัปดาห์ที่ระบุ
    ?week=YYYY-MM-DD (ถ้าไม่ระบุจะใช้สัปดาห์ปัจจุบัน)
    รองรับทั้ง schema ใหม่ (tasks[]) และเก่า (days{})
    """
    today = date_mod.today().isoformat()
    week_start = _get_monday(week if week else today)
    user_id = current_user.get("user_id", "")
    plan_id = f"{user_id}_{week_start}"
    db = get_db()
    doc = db.collection("weekly_plans").document(plan_id).get()
    if not doc.exists:
        return {"id": plan_id, "week_start": week_start, "tasks": []}

    data = doc.to_dict()
    #? backward compat: doc เก่ามี field 'days' → แปลงเป็น tasks
    if "days" in data and "tasks" not in data:
        data["tasks"] = _days_to_tasks(data.pop("days"))
    elif "days" in data:
        data.pop("days")  # ถ้ามีทั้งสองให้ใช้ tasks
    return {"id": doc.id, **data}


# ─────────────────────────────────────────────────────────────────
# POST /api/plans  — สร้าง/แทนที่แผนงาน
# ─────────────────────────────────────────────────────────────────
@router.post("/", status_code=201)
def save_plan(
    data: WeeklyPlanCreate,
    current_user: dict = Depends(get_current_user)
):
    """สร้างหรือแทนที่แผนงานเชิงพัฒนาสำหรับสัปดาห์ที่ระบุ"""
    week_start = _get_monday(data.week_start)
    user_id = current_user.get("user_id", "")
    plan_id = f"{user_id}_{week_start}"
    db = get_db()
    doc_ref = db.collection("weekly_plans").document(plan_id)

    now = _now_bkk()
    existing = doc_ref.get()
    created_at = existing.to_dict().get("created_at", now) if existing.exists else now

    #? สร้าง lookup map จากงานเดิม (id → task dict) เพื่อ preserve approval state
    old_tasks_map: dict[int, dict] = {}
    if existing.exists:
        old_data = existing.to_dict()
        old_list = old_data.get("tasks", [])
        if not old_list and "days" in old_data:
            #? backward compat: doc เก่าใช้ days dict
            old_list = _days_to_tasks(old_data["days"])
        for ot in old_list:
            old_tasks_map[ot.get("id")] = ot

    valid_dates = set(_get_week_dates(week_start))
    clean_tasks = []
    for t in data.tasks:
        task_dict = t.model_dump() if hasattr(t, 'model_dump') else dict(t)
        tid = task_dict.get("id")

        #? กรอง active_days ให้อยู่ในสัปดาห์นี้เท่านั้น
        active_days = [d for d in task_dict.get("active_days", []) if d in valid_dates]

        #? preserve approved/approved_by/approved_at จาก Firestore — ห้าม client เขียนทับ
        old = old_tasks_map.get(tid, {})
        clean_tasks.append({
            "id":          tid,
            "title":       task_dict.get("title", ""),
            "active_days": active_days,
            "description": task_dict.get("description", ""),
            "goal":        task_dict.get("goal", ""),
            "output":      task_dict.get("output", ""),
            "kpi_name":    task_dict.get("kpi_name", ""),
            "kpi_target":  task_dict.get("kpi_target", ""),
            "approved":    old.get("approved", False),
            "approved_by": old.get("approved_by", ""),
            "approved_at": old.get("approved_at", ""),
        })

    payload = {
        "user_id":    user_id,
        "user_name":  current_user.get("name", ""),
        "department": current_user.get("department", ""),
        "week_start": week_start,
        "created_at": created_at,
        "updated_at": now,
        "tasks":      clean_tasks,
    }
    doc_ref.set(payload)
    return {"id": plan_id, **payload}


# ─────────────────────────────────────────────────────────────────
# GET /api/plans/tasks  — ดึงงานในแผนสำหรับวันที่ระบุ (auto-inject)
# ─────────────────────────────────────────────────────────────────
@router.get("/tasks")
def get_plan_tasks_for_date(
    date: str,
    current_user: dict = Depends(get_current_user)
):
    """
    คืน list ของงานที่วางแผนไว้สำหรับวันที่ระบุ
    ใช้โดย emp.js สำหรับ auto-inject งานเข้าฟอร์มรายงานประจำวัน
    เฉพาะงานที่ approved=True และ date อยู่ใน active_days
    """
    week_start = _get_monday(date)
    #? ดึง user_id (personal_id) มาใช้งาน พร้อมจัดการช่องว่างที่อาจติดมา
    user_id = str(current_user.get("user_id", "")).strip()
    db = get_db()

    #? Fallback: หาก user_id ใน JWT ว่าง ให้ดึงจาก Firestore โดยตรงผ่าน email
    if not user_id:
        email = current_user.get("sub")
        if email:
            u_doc = db.collection("users").document(email).get()
            if u_doc.exists:
                user_id = str(u_doc.to_dict().get("personal_id", "")).strip()

    if not user_id:
        return []

    plan_id = f"{user_id}_{week_start}"
    doc = db.collection("weekly_plans").document(plan_id).get()
    if not doc.exists:
        return []

    doc_data = doc.to_dict()
    tasks = doc_data.get("tasks", [])

    #? backward compat: doc เก่าใช้ days dict
    if not tasks and "days" in doc_data:
        tasks = _days_to_tasks(doc_data["days"])

    #? กรองงาน: approved=True และ date อยู่ใน active_days
    #! งาน Pending และงานที่ถูกปฏิเสธ (หรือไม่มีวันนี้ใน active_days) จะไม่ถูก inject
    return [t for t in tasks if t.get("approved") is True and date in t.get("active_days", [])]


# ─────────────────────────────────────────────────────────────────
# GET /api/plans/subordinates  — หัวหน้าดูแผนของลูกน้อง
# ─────────────────────────────────────────────────────────────────
@router.get("/subordinates")
def get_subordinates_plans(
    week: str | None = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Admin/Supervisor ดูแผนงานของลูกน้องทุกคนสำหรับสัปดาห์ที่ระบุ
    Super Admin เห็นทุกคน, Supervisor (level 1-3) เห็นเฉพาะลูกน้องตามสาย
    """
    level = current_user.get("level", 0)
    role = current_user.get("role", "").lower()
    if level == 0 and 'admin' not in role:
        raise HTTPException(status_code=403, detail="สิทธิ์การเข้าถึงสำหรับผู้ดูแลระบบเท่านั้น")

    today = date_mod.today().isoformat()
    week_start = _get_monday(week if week else today)
    db = get_db()
    allowed_ids = _get_subordinate_user_ids(db, current_user)

    plans_ref = db.collection("weekly_plans").where("week_start", "==", week_start)
    result = []
    for doc in plans_ref.stream():
        d = doc.to_dict()
        uid = d.get("user_id", "")
        if allowed_ids is not None and uid not in allowed_ids:
            continue
        #? backward compat: แปลง days → tasks ถ้ายังเป็น schema เก่า
        if "days" in d and "tasks" not in d:
            d["tasks"] = _days_to_tasks(d.pop("days"))
        elif "days" in d:
            d.pop("days")
        result.append({"id": doc.id, **d})

    #? batch read: เช็ก in_report สำหรับทุก active_day ของทุก task พร้อมกัน
    all_report_refs: list = []
    ref_to_info: dict[str, tuple[str, int]] = {}  # path → (uid, task_id)

    for plan_dict in result:
        uid = plan_dict.get("user_id", "")
        for task in plan_dict.get("tasks", []):
            tid = task.get("id")
            for date_str in task.get("active_days", []):
                report_id = f"{uid}_{date_str}"
                ref = db.collection("reports").document(report_id)
                all_report_refs.append(ref)
                ref_to_info[ref.path] = (uid, tid)

    #? ดึงทุก report doc พร้อมกัน (1 round-trip)
    exists_set: set[str] = set()
    if all_report_refs:
        for snap in db.get_all(all_report_refs):
            if snap.exists:
                exists_set.add(snap.reference.path)

    #? สร้าง set (uid, task_id) ที่มี report อย่างน้อย 1 วัน
    in_report_keys: set[tuple[str, int]] = set()
    for path, (uid, tid) in ref_to_info.items():
        if path in exists_set:
            in_report_keys.add((uid, tid))

    #? ใส่ in_report flag ให้แต่ละ task
    for plan_dict in result:
        uid = plan_dict.get("user_id", "")
        for task in plan_dict.get("tasks", []):
            task["in_report"] = (uid, task.get("id")) in in_report_keys

    return result


# ─────────────────────────────────────────────────────────────────
# PATCH /api/plans/{plan_id}/approve  — อนุมัติ/ยกเลิกงานเดี่ยว
# ─────────────────────────────────────────────────────────────────
@router.patch("/{plan_id}/approve")
def approve_task(
    plan_id: str,
    data: TaskApprovalUpdate,
    current_user: dict = Depends(get_current_user)
):
    """
    Supervisor อนุมัติหรือยกเลิกการอนุมัติงานแต่ละรายการในแผนงานเชิงพัฒนา
    """
    level = current_user.get("level", 0)
    role = current_user.get("role", "").lower()
    if level == 0 and 'admin' not in role:
        raise HTTPException(status_code=403, detail="สิทธิ์การเข้าถึงสำหรับผู้ดูแลระบบเท่านั้น")

    db = get_db()
    doc_ref = db.collection("weekly_plans").document(plan_id)
    doc = doc_ref.get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="ไม่พบแผนงานที่ระบุ")

    plan_data = doc.to_dict()
    tasks = plan_data.get("tasks", [])

    #? backward compat
    if not tasks and "days" in plan_data:
        tasks = _days_to_tasks(plan_data["days"])

    #? หา task ที่ต้องการในรายการ tasks (ระดับสัปดาห์)
    target_task = next((t for t in tasks if t.get("id") == data.task_id), None)
    if target_task is None:
        raise HTTPException(status_code=404, detail="ไม่พบงานที่ระบุในแผนงาน")

    #! ป้องกันการยกเลิกอนุมัติ หากงานนี้มีรายงานประจำวันแล้วใน active_days ใดก็ตาม
    if not data.approved:
        uid = plan_data.get("user_id", "")
        for date_str in target_task.get("active_days", []):
            report_id = f"{uid}_{date_str}"
            if db.collection("reports").document(report_id).get().exists:
                raise HTTPException(
                    status_code=409,
                    detail="ไม่สามารถยกเลิกการอนุมัติได้ เนื่องจากงานนี้อยู่ระหว่างดำเนินการในรายงานแล้ว"
                )

    target_task["approved"]    = data.approved
    target_task["approved_by"] = current_user.get("name", current_user.get("sub", ""))
    target_task["approved_at"] = _now_bkk() if data.approved else ""

    doc_ref.update({"tasks": tasks, "updated_at": _now_bkk()})
    return {"status": "ok", "plan_id": plan_id, "task_id": data.task_id, "approved": data.approved}


# ─────────────────────────────────────────────────────────────────
# Helpers สำหรับ Excel Export แผนงาน
# ─────────────────────────────────────────────────────────────────
_TH_DAYS = ["จันทร์", "อังคาร", "พุธ", "พฤหัสบดี", "ศุกร์", "เสาร์", "อาทิตย์"]
_TH_DAYS_SHORT = ["จ", "อ", "พ", "พฤ", "ศ", "ส"]
_TH_MON_SHORT = ["", "ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
                 "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
_TH_MON_LONG  = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
                 "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]


def _format_date_th(date_str: str) -> str:
    """แปลง YYYY-MM-DD → "วัน DD เดือน" ภาษาไทย"""
    d = date_mod.fromisoformat(date_str)
    return f"{_TH_DAYS[d.weekday()]} {d.day} {_TH_MON_SHORT[d.month]}"


def _format_active_days_th(active_days: list[str]) -> str:
    """แปลง list ของ YYYY-MM-DD → "จ, อ, พ" ย่อวัน"""
    result = []
    for ds in sorted(active_days):
        try:
            d = date_mod.fromisoformat(ds)
            result.append(_TH_DAYS_SHORT[d.weekday()])
        except Exception:
            pass
    return ", ".join(result)


def _get_months_mondays(month: str) -> list[str]:
    """คืน list ของวันจันทร์ (week_start) ทุกสัปดาห์ที่มีวันอยู่ในเดือน YYYY-MM"""
    import calendar as _cal
    year, mon = int(month[:4]), int(month[5:7])
    first_day = date_mod(year, mon, 1)
    last_day  = date_mod(year, mon, _cal.monthrange(year, mon)[1])
    start_monday = first_day - timedelta(days=first_day.weekday())
    mondays, d = [], start_monday
    while d <= last_day:
        mondays.append(d.isoformat())
        d += timedelta(weeks=1)
    return mondays


def _build_plans_wb(rows: list[dict], title: str, has_week_col: bool):
    """
    สร้าง openpyxl Workbook สำหรับ Excel แผนงานเชิงพัฒนา
    rows แต่ละ dict: user_id, user_name, department, week_label?, task_num,
                    active_days_str, title, goal, output, kpi_name, kpi_target,
                    approved, approved_by
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    def side():
        return Side(style="thin", color="CCCCCC")
    def bdr():
        return Border(left=side(), right=side(), top=side(), bottom=side())

    hdr_fill   = PatternFill("solid", fgColor="1059A3")
    emp_fill   = PatternFill("solid", fgColor="E8EFF8")
    green_fill = PatternFill("solid", fgColor="D4EDDA")
    orng_fill  = PatternFill("solid", fgColor="FFF3CD")
    red_fill   = PatternFill("solid", fgColor="F8D7DA")
    alt_fill   = PatternFill("solid", fgColor="F9FAFB")

    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    emp_font  = Font(bold=True, color="1A3A6B", size=10)
    body_font = Font(size=10)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_va   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    if has_week_col:
        headers    = ["ลำดับ", "ชื่อ-สกุล", "แผนก", "สัปดาห์ที่", "วันที่ทำงาน", "งานที่",
                      "ชื่องาน", "เป้าหมาย", "ผลผลิต", "ชื่อ KPI", "เป้าหมาย KPI", "สถานะ"]
        col_widths = [6, 18, 14, 12, 16, 6, 24, 20, 20, 18, 14, 14]
        center_cols = {1, 6}
    else:
        headers    = ["ลำดับ", "ชื่อ-สกุล", "แผนก", "วันที่ทำงาน", "งานที่",
                      "ชื่องาน", "เป้าหมาย", "ผลผลิต", "ชื่อ KPI", "เป้าหมาย KPI", "สถานะ"]
        col_widths = [6, 18, 14, 16, 6, 24, 20, 20, 18, 14, 14]
        center_cols = {1, 5}

    COLS = len(headers)

    def col_letter(n: int) -> str:
        return chr(ord('A') + n - 1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Title (row 1)
    ws.merge_cells(f"A1:{col_letter(COLS)}1")
    c = ws["A1"]
    c.value     = title
    c.font      = Font(bold=True, size=12, color="1A3A6B")
    c.alignment = center
    ws.row_dimensions[1].height = 28

    # Header (row 2)
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c            = ws.cell(row=2, column=ci, value=h)
        c.font       = hdr_font
        c.fill       = hdr_fill
        c.alignment  = center
        c.border     = bdr()
        ws.column_dimensions[col_letter(ci)].width = w
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    row, seq, last_emp = 3, 1, None

    for entry in rows:
        uid = entry.get("user_id", "")

        # Employee separator
        if uid != last_emp:
            ws.merge_cells(f"A{row}:{col_letter(COLS)}{row}")
            c            = ws.cell(row=row, column=1)
            c.value      = f"  {entry['user_name']}  [{entry['department']}]"
            c.font       = emp_font
            c.fill       = emp_fill
            c.alignment  = left_va
            for ci in range(1, COLS + 1):
                ws.cell(row=row, column=ci).border = bdr()
            ws.row_dimensions[row].height = 22
            row      += 1
            last_emp  = uid

        # Status
        if entry.get("approved"):
            s_fill, status_txt = green_fill, "✓ อนุมัติแล้ว"
        elif entry.get("approved_by"):
            s_fill, status_txt = red_fill,   "✗ ไม่อนุมัติ"
        else:
            s_fill, status_txt = orng_fill,  "⋯ รอการอนุมัติ"

        row_fill = PatternFill("solid", fgColor="FFFFFF") if row % 2 == 0 else alt_fill

        if has_week_col:
            values = [seq, entry["user_name"], entry["department"], entry.get("week_label", ""),
                      entry.get("active_days_str", ""), entry["task_num"],
                      entry["title"], entry["goal"], entry["output"],
                      entry["kpi_name"], entry["kpi_target"], status_txt]
        else:
            values = [seq, entry["user_name"], entry["department"],
                      entry.get("active_days_str", ""), entry["task_num"],
                      entry["title"], entry["goal"], entry["output"],
                      entry["kpi_name"], entry["kpi_target"], status_txt]

        for ci, val in enumerate(values, 1):
            c            = ws.cell(row=row, column=ci, value=val)
            c.font       = body_font
            c.fill       = s_fill if ci == COLS else row_fill
            c.alignment  = center if ci in center_cols else left_va
            c.border     = bdr()
        ws.row_dimensions[row].height = 18
        seq += 1
        row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─────────────────────────────────────────────────────────────────
# GET /api/plans/export/weekly  — ส่งออก Excel แผนงานรายสัปดาห์
# ─────────────────────────────────────────────────────────────────
@router.get("/export/weekly")
def export_plans_weekly(
    week: str | None = None,
    current_user: dict = Depends(get_current_user)
):
    """Export แผนงานเชิงพัฒนาของลูกน้องทั้งหมดในสัปดาห์ที่ระบุ เป็น Excel (.xlsx)"""
    level = current_user.get("level", 0)
    role  = current_user.get("role", "").lower()
    if level == 0 and 'admin' not in role:
        raise HTTPException(status_code=403, detail="สิทธิ์การเข้าถึงสำหรับผู้ดูแลระบบเท่านั้น")

    today      = date_mod.today().isoformat()
    week_start = _get_monday(week if week else today)
    db         = get_db()
    allowed_ids = _get_subordinate_user_ids(db, current_user)

    plans = []
    for doc in db.collection("weekly_plans").where("week_start", "==", week_start).stream():
        d = doc.to_dict()
        if allowed_ids is not None and d.get("user_id", "") not in allowed_ids:
            continue
        if "days" in d and "tasks" not in d:
            d["tasks"] = _days_to_tasks(d.pop("days"))
        elif "days" in d:
            d.pop("days")
        plans.append({"id": doc.id, **d})
    plans.sort(key=lambda p: (p.get("department", ""), p.get("user_name", "")))

    rows = []
    for ti_plan, plan in enumerate(plans):
        for ti, task in enumerate(plan.get("tasks", []), 1):
            rows.append({
                "user_id":        plan.get("user_id", ""),
                "user_name":      plan.get("user_name", ""),
                "department":     plan.get("department", ""),
                "task_num":       ti,
                "active_days_str": _format_active_days_th(task.get("active_days", [])),
                "title":          task.get("title", ""),
                "goal":           task.get("goal", ""),
                "output":         task.get("output", ""),
                "kpi_name":       task.get("kpi_name", ""),
                "kpi_target":     task.get("kpi_target", ""),
                "approved":       task.get("approved", False),
                "approved_by":    task.get("approved_by", ""),
            })

    week_dates = _get_week_dates(week_start)
    s = date_mod.fromisoformat(week_dates[0])
    e = date_mod.fromisoformat(week_dates[5])
    title = (f"แผนงานเชิงพัฒนารายสัปดาห์  {s.day} {_TH_MON_SHORT[s.month]}"
             f" — {e.day} {_TH_MON_SHORT[e.month]} {e.year + 543}")

    output = _build_plans_wb(rows, title, has_week_col=False)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="plans_weekly_{week_start}.xlsx"'},
    )


# ─────────────────────────────────────────────────────────────────
# GET /api/plans/export/monthly  — ส่งออก Excel แผนงานรายเดือน
# ─────────────────────────────────────────────────────────────────
@router.get("/export/monthly")
def export_plans_monthly(
    month: str,
    current_user: dict = Depends(get_current_user)
):
    """Export แผนงานเชิงพัฒนาของลูกน้องทั้งหมดในเดือนที่ระบุ (YYYY-MM) เป็น Excel (.xlsx)"""
    level = current_user.get("level", 0)
    role  = current_user.get("role", "").lower()
    if level == 0 and 'admin' not in role:
        raise HTTPException(status_code=403, detail="สิทธิ์การเข้าถึงสำหรับผู้ดูแลระบบเท่านั้น")

    if not month or len(month) < 7:
        raise HTTPException(status_code=422, detail="month ต้องอยู่ในรูปแบบ YYYY-MM")

    db          = get_db()
    allowed_ids = _get_subordinate_user_ids(db, current_user)
    mondays     = _get_months_mondays(month)

    plans_by_week: dict[str, list] = {}
    for week_start in mondays:
        week_plans = []
        for doc in db.collection("weekly_plans").where("week_start", "==", week_start).stream():
            d = doc.to_dict()
            if allowed_ids is not None and d.get("user_id", "") not in allowed_ids:
                continue
            if "days" in d and "tasks" not in d:
                d["tasks"] = _days_to_tasks(d.pop("days"))
            elif "days" in d:
                d.pop("days")
            week_plans.append({"id": doc.id, **d})
        plans_by_week[week_start] = week_plans

    emp_map: dict[str, dict] = {}
    for week_plans in plans_by_week.values():
        for p in week_plans:
            uid = p.get("user_id", "")
            if uid and uid not in emp_map:
                emp_map[uid] = {"user_name": p.get("user_name", ""),
                                "department": p.get("department", "")}
    sorted_emps = sorted(emp_map.items(),
                         key=lambda x: (x[1].get("department", ""), x[1].get("user_name", "")))

    rows = []
    for uid, emp_info in sorted_emps:
        for wi, week_start in enumerate(mondays, 1):
            plan = next((p for p in plans_by_week.get(week_start, [])
                         if p.get("user_id") == uid), None)
            if not plan:
                continue
            for ti, task in enumerate(plan.get("tasks", []), 1):
                rows.append({
                    "user_id":        uid,
                    "user_name":      emp_info["user_name"],
                    "department":     emp_info["department"],
                    "week_label":     f"สัปดาห์ที่ {wi}",
                    "task_num":       ti,
                    "active_days_str": _format_active_days_th(task.get("active_days", [])),
                    "title":          task.get("title", ""),
                    "goal":           task.get("goal", ""),
                    "output":         task.get("output", ""),
                    "kpi_name":       task.get("kpi_name", ""),
                    "kpi_target":     task.get("kpi_target", ""),
                    "approved":       task.get("approved", False),
                    "approved_by":    task.get("approved_by", ""),
                })

    year_val, mon_val = int(month[:4]), int(month[5:7])
    title = f"แผนงานเชิงพัฒนารายเดือน  {_TH_MON_LONG[mon_val]} {year_val + 543}"

    output = _build_plans_wb(rows, title, has_week_col=True)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="plans_monthly_{month}.xlsx"'},
    )
