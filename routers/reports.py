from fastapi import APIRouter, HTTPException, Body, Depends, Request
from fastapi.responses import StreamingResponse
from database import get_db
from models import ReportCreate, ReportOut, CommentCreate, CommentModel, TaskModel
from auth import get_current_user
from activity_logger import log_activity, LogAction
from typing import List, Optional
import uuid
import io
from datetime import datetime
from zoneinfo import ZoneInfo

#? กำหนด Router สำหรับจัดการรายงานประจำวัน (Reports)
router = APIRouter()
#? ตั้งค่า Timezone เป็นประเทศไทยเพื่อให้เวลาในรายงานถูกต้อง
bz_tz = ZoneInfo('Asia/Bangkok')

#? API สำหรับส่งรายงานประจำวัน (หากส่งซ้ำในวันเดียวกันจะถือเป็นการแก้ไขรายงานเดิม)
@router.post("/", response_model=ReportOut)
def create_report(report: ReportCreate, request: Request, current_user: dict = Depends(get_current_user)):
    db = get_db()

    now = datetime.now(bz_tz)
    #? ดึงวันที่ปัจจุบันมาใช้ตั้งชื่อ ID (YYYY-MM-DD)
    date_str = now.strftime('%Y-%m-%d')
    #? บันทึกวันเวลาที่ส่งรายงานแบบเต็ม (ใช้แสดงผลในตาราง)
    timestamp = now.strftime('%Y-%m-%d %H:%M:%S')
    #? แยกเฉพาะเวลาที่ส่งเพื่อใช้แสดงผลในแดชบอร์ด
    submit_time = now.strftime('%H:%M')

    #? ใช้ user_id ต่อด้วยวันที่เป็น ID ของ Document เพื่อป้องกันการส่งรายงานซ้ำซ้อนในวันเดียวกัน
    report_id = f"{report.user_id}_{date_str}"

    report_dict = report.model_dump()
    report_dict['id'] = report_id
    report_dict['timestamp'] = timestamp
    report_dict['submit_time'] = submit_time

    #? ตรวจสอบว่าพนักงานคนนี้เคยส่งรายงานของวันนี้มาแล้วหรือยัง
    doc = db.collection('reports').document(report_id).get()
    if doc.exists:
        #? หากเคยส่งแล้ว ให้ดึงคอมเมนท์เดิมที่มีอยู่กลับมา (เพื่อไม่ให้คอมเมนท์หายไปเมื่อแก้ไขรายงาน)
        report_dict['comments'] = doc.to_dict().get('comments', [])
    else:
        #? หากเป็นการส่งครั้งแรกของวัน ให้กำหนดรายการคอมเมนท์เป็นอาเรย์ว่าง
        report_dict['comments'] = []

    db.collection('reports').document(report_id).set(report_dict)

    log_activity(db, action=LogAction.REPORT_SUBMIT, request=request, user=current_user,
                 resource_id=report_id, resource_type="report",
                 details={"date": date_str, "work_mode": report.work_mode, "progress": report.progress})

    return report_dict

#? API ดึงรายการรายงานทั้งหมด (คัดกรองตามสิทธิ์การเข้าถึง)
@router.get("/", response_model=List[ReportOut])
def get_reports(date: str = None, current_user: dict = Depends(get_current_user)):
    level = current_user.get("level", 0)
    role = current_user.get("role", "").lower()
    personal_id = current_user.get("user_id")

    #? ตรวจสอบประเภทผู้ใช้งาน: Super Admin (Level 9 หรือมี Role 'admin') 
    #? จะมีสิทธิ์สูงสุดในการเข้าถึงข้อมูลพนักงานทุกคนได้โดยไม่ต้องกรอง
    is_super_admin = level == 9 or 'admin' in role

    db = get_db()

    #? ตัวแปรสำหรับเก็บรายชื่อพนักงานที่ Supervisor ท่านนี้มีสิทธิ์ประเมิน/มองเห็น
    allowed_target_ids = set()
    if not is_super_admin and 1 <= level <= 3:
        #? ค้นหาสายการบังคับบัญชาจาก 'evaluations' โดยเปรียบเทียบจาก personal_id ของหัวหน้า
        #? เพื่อดึงรายชื่อเฉพาะพนักงาน (Target) ที่อยู่ในความดูแลเท่านั้นมาแสดงผล
        for e in db.collection("evaluations") \
                    .where("evaluator_ids", "array_contains", personal_id) \
                    .stream():
            allowed_target_ids.add(e.to_dict().get("target_id"))

    reports_ref = db.collection('reports')
    #todo ควรเปลี่ยนไปใช้ .where() ในการ Query ตั้งแต่แรกแทนการดึงมาทั้งหมดเพื่อประหยัด Resource
    #! การใช้ .stream() ดึงข้อมูลทั้งหมดมาวนลูปกรองใน Python จะทำงานช้าลงหากข้อมูลมีจำนวนมหาศาล
    docs = reports_ref.stream()

    reports = []
    for doc in docs:
        data = doc.to_dict()
        if date and not data.get('timestamp', '').startswith(date):
            continue

        r_user_id = data.get('user_id')

        #? กรองข้อมูลแยกตามลำดับขั้นความปลอดภัย
        if is_super_admin:
            pass  #? สิทธิ์สูงสุด: ไม่ต้องกรองข้าม (มองเห็นได้ทุกคน)
        elif level == 0:
            if r_user_id != personal_id:
                #? พนักงานระดับทั่วไป: มองเห็นได้เฉพาะรายงานของตัวเองเท่านั้น
                continue
        elif 1 <= level <= 3:
            if r_user_id not in allowed_target_ids and r_user_id != personal_id:
                #? หัวหน้างาน: มองเห็นรายงานของตัวเองและลูกน้องในสายงานที่ได้รับอนุญาตเท่านั้น
                continue
        else:
            if r_user_id != personal_id:
                #? กรณีอื่นๆ ที่ไม่ได้ระบุไว้: ให้ปลอดภัยไว้ก่อนโดยให้เห็นได้เฉพาะตัวเอง
                continue

        reports.append(data)

    reports.sort(key=lambda x: x.get('timestamp', ''), reverse=True)
    return reports

@router.get("/{report_id}", response_model=ReportOut)
def get_report(report_id: str, request: Request, current_user: dict = Depends(get_current_user)):
    db = get_db()
    doc = db.collection('reports').document(report_id).get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Report not found")

    #? log เฉพาะเมื่อ supervisor (level >= 1) เปิดดูรายงานของพนักงาน
    level = current_user.get("level", 0)
    if level >= 1:
        log_activity(db, action=LogAction.REPORT_VIEW, request=request, user=current_user,
                     resource_id=report_id, resource_type="report")

    return doc.to_dict()

@router.patch("/{report_id}/tasks")
def update_tasks(report_id: str, tasks: List[TaskModel] = Body(...),
                 source: Optional[str] = None,
                 request: Request = None,
                 current_user: dict = Depends(get_current_user)):
    #? source เป็น query param — ส่ง ?source=manual เมื่อ user กด save เอง (ไม่ใช่ auto-save)
    db = get_db()
    report_ref = db.collection('reports').document(report_id)
    doc = report_ref.get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Report not found")
    tasks_data = [t.model_dump() for t in tasks]
    #? ทำการอัปเดตเฉพาะฟิลด์ 'tasks' ในเอกสารรายงานเดิม
    report_ref.update({'tasks': tasks_data})

    #? log เฉพาะ manual save — ไม่ log auto-save ทุก 30 วินาที
    if source == "manual":
        log_activity(db, action=LogAction.REPORT_TASK_UPDATE, request=request, user=current_user,
                     resource_id=report_id, resource_type="report",
                     details={"task_count": len(tasks_data)})

    return {"success": True}

#? เพิ่มคอมเมนท์ลงในรายงาน
@router.post("/{report_id}/comments", response_model=CommentModel)
def add_comment(report_id: str, comment: CommentCreate, request: Request, current_user: dict = Depends(get_current_user)):
    db = get_db()
    report_ref = db.collection('reports').document(report_id)
    doc = report_ref.get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Report not found")

    #? สร้างรหัสสุ่ม ID (UUID) สำหรับอ้างอิงและจัดการคอมเมนท์แต่ละรายการในภายหลัง
    comment_id = str(uuid.uuid4())
    now = datetime.now(bz_tz)
    #? บันทึกเฉพาะเวลาที่ส่งคอมเมนท์ (เช่น 15:45) เพื่อความกระชับในการแสดงผลในแชท
    timestamp_str = now.strftime('%H:%M')

    #? เตรียมข้อมูลคอมเมนท์ในรูปแบบ Dictionary เพื่อบันทึกลง Firestore
    comment_dict = comment.model_dump()
    comment_dict['id'] = comment_id
    comment_dict['timestamp'] = timestamp_str

    report_data = doc.to_dict()
    comments = report_data.get('comments', [])
    comments.append(comment_dict)

    report_ref.update({'comments': comments})

    log_activity(db, action=LogAction.REPORT_COMMENT_ADD, request=request, user=current_user,
                 resource_id=report_id, resource_type="report",
                 details={"tag": comment.tag or "", "author": comment.author_name or ""})

    return comment_dict


#? Export รายงานประจำวันเป็นไฟล์ Excel (เจ้าของรายงาน หรือ supervisor/admin เท่านั้น)
@router.get("/{report_id}/export")
def export_report_excel(report_id: str, current_user: dict = Depends(get_current_user)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    db = get_db()
    doc = db.collection('reports').document(report_id).get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Report not found")

    data = doc.to_dict()
    level = current_user.get("level", 0)
    role = current_user.get("role", "").lower()
    is_admin = level == 9 or 'admin' in role

    #! ตรวจสิทธิ์: เจ้าของรายงาน หรือ supervisor (level >= 1) หรือ admin
    if not is_admin and level < 1 and data.get("user_id") != current_user.get("user_id"):
        raise HTTPException(status_code=403, detail="Access denied")

    # ── helper styles ─────────────────────────────────────────────────
    def _side():
        return Side(style="thin", color="CCCCCC")

    def _border():
        return Border(left=_side(), right=_side(), top=_side(), bottom=_side())

    HDR_FILL   = PatternFill("solid", fgColor="1059A3")
    INFO_FILL  = PatternFill("solid", fgColor="E8EFF8")
    SEC_FILL   = PatternFill("solid", fgColor="EBF2FB")
    DONE_FILL  = PatternFill("solid", fgColor="D4EDDA")
    PROG_FILL  = PatternFill("solid", fgColor="FFF3CD")
    PEND_FILL  = PatternFill("solid", fgColor="F9FAFB")
    ALT_FILL   = PatternFill("solid", fgColor="F9FAFB")

    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    sec_font  = Font(bold=True, color="1A3A6B", size=10)
    body_font = Font(size=10)
    wrap_c    = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    center_c  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_c    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    STATUS_LABEL = {"done": "✓ เสร็จแล้ว", "prog": "⋯ กำลังดำเนินการ", "pend": "◯ ยังไม่เริ่ม"}
    MODE_LABEL   = {"wfh": "Work from Home", "onsite": "On-site", "hybrid": "Hybrid"}

    def _fmt_elapsed(secs):
        if secs is None:
            return "-"
        secs = int(secs)
        h, rem = divmod(secs, 3600)
        m, s   = divmod(rem, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"

    def _set(ws, row, col, val, font=None, fill=None, align=None, border=True):
        c = ws.cell(row=row, column=col, value=val)
        if font:   c.font      = font
        if fill:   c.fill      = fill
        if align:  c.alignment = align
        if border: c.border    = _border()
        return c

    # ── workbook ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายงานประจำวัน"

    COLS = 6
    col_letters = [ws.cell(row=1, column=i+1).column_letter for i in range(COLS)]

    # Row 1: Title
    ws.merge_cells(f"A1:{col_letters[-1]}1")
    c = _set(ws, 1, 1, "สรุปรายงานการทำงานประจำวัน", font=Font(bold=True, color="FFFFFF", size=13), fill=HDR_FILL, align=center_c)
    ws.row_dimensions[1].height = 36

    # Row 2: employee info labels
    info_labels = ["ชื่อ-นามสกุล", "ตำแหน่ง", "หน่วยงาน", "วันที่ส่ง", "รูปแบบงาน", "ความคืบหน้า"]
    for i, lbl in enumerate(info_labels):
        _set(ws, 2, i+1, lbl, font=Font(bold=True, color="1A3A6B", size=9), fill=INFO_FILL, align=center_c)
    ws.row_dimensions[2].height = 20

    # Row 3: employee info values
    wmode = MODE_LABEL.get(data.get("work_mode", ""), data.get("work_mode", ""))
    info_vals = [
        data.get("name", ""),
        data.get("role", ""),
        data.get("department", ""),
        data.get("timestamp", ""),
        wmode,
        f"{data.get('progress', 0)}%",
    ]
    for i, val in enumerate(info_vals):
        _set(ws, 3, i+1, val, font=body_font, fill=PatternFill("solid", fgColor="F0F6FF"), align=center_c)
    ws.row_dimensions[3].height = 20

    # Row 4: blank spacer
    ws.row_dimensions[4].height = 8

    # Row 5: section header — รายการงาน
    ws.merge_cells(f"A5:{col_letters[-1]}5")
    _set(ws, 5, 1, "รายการงาน", font=sec_font, fill=SEC_FILL, align=left_c)
    ws.row_dimensions[5].height = 22

    # Row 6: task column headers
    task_headers = ["#", "ชื่องาน", "ประเภท", "สถานะ", "เวลาที่ใช้", "รายละเอียด"]
    for i, h in enumerate(task_headers):
        _set(ws, 6, i+1, h, font=Font(bold=True, color="FFFFFF", size=10), fill=HDR_FILL, align=center_c)
    ws.row_dimensions[6].height = 22

    # Task rows
    tasks = data.get("tasks", [])
    row = 7
    for idx, task in enumerate(tasks):
        status = task.get("status", "pend")
        task_fill = DONE_FILL if status == "done" else (PROG_FILL if status == "prog" else PEND_FILL)
        if idx % 2 == 1 and status == "pend":
            task_fill = ALT_FILL

        elapsed = task.get("elapsed_seconds")
        #? ถ้า task กำลัง active (started_at มีค่า elapsed_seconds เป็น 0 หรือ None)
        if task.get("started_at") and status == "prog":
            elapsed = elapsed  # ใช้ค่าที่บันทึกล่าสุด

        row_vals = [
            idx + 1,
            task.get("title", ""),
            task.get("task_type", ""),
            STATUS_LABEL.get(status, status),
            _fmt_elapsed(elapsed),
            task.get("description", "") or "",
        ]
        col_aligns = [center_c, left_c, center_c, center_c, center_c, wrap_c]
        for i, (val, aln) in enumerate(zip(row_vals, col_aligns)):
            _set(ws, row, i+1, val, font=body_font, fill=task_fill, align=aln)
        ws.row_dimensions[row].height = 28
        row += 1

    if not tasks:
        ws.merge_cells(f"A{row}:{col_letters[-1]}{row}")
        _set(ws, row, 1, "ไม่มีรายการงาน", font=Font(color="999999", size=10), fill=PEND_FILL, align=center_c)
        row += 1

    # Blank spacer
    row += 1

    # Section: ปัญหาที่พบ
    ws.merge_cells(f"A{row}:{col_letters[-1]}{row}")
    _set(ws, row, 1, "ปัญหาที่พบ", font=sec_font, fill=SEC_FILL, align=left_c)
    ws.row_dimensions[row].height = 22
    row += 1
    ws.merge_cells(f"A{row}:{col_letters[-1]}{row}")
    problems_text = data.get("problems", "") or "-"
    _set(ws, row, 1, problems_text, font=body_font, align=wrap_c)
    ws.row_dimensions[row].height = max(40, len(problems_text) // 6 * 14 + 20)
    row += 1

    # Blank spacer
    row += 1

    # Section: แผนพรุ่งนี้
    ws.merge_cells(f"A{row}:{col_letters[-1]}{row}")
    _set(ws, row, 1, "แผนสำหรับวันพรุ่งนี้", font=sec_font, fill=SEC_FILL, align=left_c)
    ws.row_dimensions[row].height = 22
    row += 1
    ws.merge_cells(f"A{row}:{col_letters[-1]}{row}")
    plan_text = data.get("plan_tomorrow", "") or "-"
    _set(ws, row, 1, plan_text, font=body_font, align=wrap_c)
    ws.row_dimensions[row].height = max(40, len(plan_text) // 6 * 14 + 20)
    row += 1

    # Section: comments (ถ้ามี)
    comments = data.get("comments", [])
    if comments:
        row += 1
        ws.merge_cells(f"A{row}:{col_letters[-1]}{row}")
        _set(ws, row, 1, "ความคิดเห็นจากหัวหน้างาน", font=sec_font, fill=SEC_FILL, align=left_c)
        ws.row_dimensions[row].height = 22
        row += 1

        cmt_headers = ["#", "ผู้เขียน", "เวลา", "แท็ก", "", "ข้อความ"]
        for i, h in enumerate(cmt_headers):
            if h:
                _set(ws, row, i+1, h, font=Font(bold=True, color="FFFFFF", size=10), fill=HDR_FILL, align=center_c)
        ws.row_dimensions[row].height = 20
        row += 1

        TAG_LABEL = {
            "acknowledge": "รับทราบ",
            "needs_fix":   "ต้องแก้ไข",
            "great":       "ดีมาก",
            "urgent":      "ติดตามด่วน",
        }
        for ci, cmt in enumerate(comments):
            cmt_fill = PatternFill("solid", fgColor="F0F6FF") if ci % 2 == 0 else PatternFill("solid", fgColor="FAFAFA")
            tag_raw = cmt.get("tag") or ""
            cmt_vals = [
                ci + 1,
                cmt.get("author_name", ""),
                cmt.get("timestamp", ""),
                TAG_LABEL.get(tag_raw, tag_raw),
                "",
                cmt.get("message", ""),
            ]
            col_aligns2 = [center_c, left_c, center_c, center_c, center_c, wrap_c]
            for i, (val, aln) in enumerate(zip(cmt_vals, col_aligns2)):
                _set(ws, row, i+1, val, font=body_font, fill=cmt_fill, align=aln)
            ws.row_dimensions[row].height = 28
            row += 1

    # Column widths
    col_widths = [5, 35, 18, 16, 12, 35]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[col_letters[i]].width = w

    ws.freeze_panes = "A4"

    # ── stream response ────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    date_part = data.get("timestamp", "")[:10] or report_id
    filename  = f"daily_report_{report_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
