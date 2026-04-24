from fastapi import APIRouter, HTTPException, Header, Request
from fastapi.responses import StreamingResponse
from database import get_db
from auth import decode_access_token
from activity_logger import log_activity, LogAction
from pydantic import BaseModel
from typing import Optional, List
import re
import calendar
import bcrypt
from datetime import date as date_mod


#? ตรวจสอบว่ารหัสผ่านนี้ผ่านการ hash ด้วย bcrypt แล้วหรือยัง
def _is_hashed(password: str) -> bool:
    return password.startswith("$2b$") or password.startswith("$2a$")


#? Hash รหัสผ่านด้วย bcrypt (rounds=12) — ใช้ก่อนบันทึกลง Firestore เสมอ
def _hash_password(plain: str) -> str:
    return bcrypt.hashpw(plain.encode("utf-8"), bcrypt.gensalt(rounds=12)).decode("utf-8")

#? กำหนด Router สำหรับระบบจัดการของ Admin
router = APIRouter()


#? ตรวจสอบสิทธิ์ว่าเป็น Super Admin (Level 9 หรือ Role เป็น Admin) หรือไม่
#! หาก Token ไม่ถูกต้องหรือไม่มีสิทธิ์ จะส่ง Error 401 หรือ 403 กลับไปทันที
def _require_super_admin(authorization: str):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = authorization.removeprefix("Bearer ")
    payload = decode_access_token(token)
    level = payload.get("level", 0)
    role = payload.get("role", "").lower()
    if not (level == 9 or 'admin' in role):
        raise HTTPException(status_code=403, detail="Forbidden: super admin only")
    return get_db(), payload


#? ตรวจสอบสิทธิ์ว่ามีสิทธิ์ระดับ Admin หรือ Supervisor (Level >= 1) หรือไม่
def _require_any_admin(authorization: str):
    """Authenticate any admin-level user (level >= 1 or role contains 'admin').
    Returns (db, user_data) tuple."""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = authorization.removeprefix("Bearer ")
    payload = decode_access_token(token)
    level = payload.get("level", 0)
    role = payload.get("role", "").lower()
    if level == 0 and 'admin' not in role:
        raise HTTPException(status_code=403, detail="Forbidden: admin access required")
    return get_db(), payload


#? คำนวณรายชื่อ User ID ที่ Admin/Supervisor คนนั้นสามารถมองเห็นได้
#! ระบบจะคืนค่า None สำหรับ Super Admin หมายความว่ามองเห็นทุกคน
def _get_visible_user_ids(db, user_data):
    """Return set of personal_ids this user can see.
    Returns None for super admin (meaning 'all users').
    user_data is a JWT payload dict — uses 'user_id' key for personal_id."""
    level = user_data.get("level", 0)
    role = user_data.get("role", "").lower()
    personal_id = user_data.get("user_id")  # JWT payload key
    is_super_admin = level == 9 or 'admin' in role
    if is_super_admin:
        return None  # None = see all
    allowed = {personal_id}  # always include self
    if 1 <= level <= 3:
        #? ค้นหาลูกน้องที่อยู่ในสายการบังคับบัญชาจาก Collection 'evaluations'
        for e in db.collection("evaluations") \
                    .where("evaluator_ids", "array_contains", personal_id) \
                    .stream():
            allowed.add(e.to_dict().get("target_id"))
    return allowed


#? โครงสร้างข้อมูลสิ่งที่ต้องส่งมาเมื่อต้องการ "สร้างพนักงานใหม่" (Required Fields)
class UserCreate(BaseModel):
    personal_id: str
    firstname: str
    lastname: str
    email: str
    position: str = ""
    department: str = ""
    agency: str = ""
    level: int = 0
    role: str = "employee"
    password: str
    ignore: int = 0


#? โครงสร้างข้อมูลเมื่อต้องการ "แก้ไขพนักงาน" (Optional Fields — เลือกส่งเฉพาะฟิลด์ที่ต้องการแก้)
class UserUpdate(BaseModel):
    personal_id: Optional[str] = None
    firstname: Optional[str] = None
    lastname: Optional[str] = None
    position: Optional[str] = None
    department: Optional[str] = None
    agency: Optional[str] = None
    level: Optional[int] = None
    role: Optional[str] = None
    password: Optional[str] = None
    ignore: Optional[int] = None


#? ดึงรายชื่อพนักงานทั้งหมดในระบบ (เฉพาะ Super Admin)
@router.get("/users")
def list_all_users(authorization: str = Header(None)):
    db, _ = _require_super_admin(authorization)
    result = []
    for doc in db.collection("users").stream():
        u = doc.to_dict()
        u.setdefault('ignore', 0)
        result.append(u)
    result.sort(key=lambda x: (x.get('department', ''), x.get('firstname', '')))
    return result


#? สร้างบัญชีผู้ใช้งานใหม่
#! ต้องตรวจสอบซ้ำว่า Email ซ้ำในระบบหรือไม่ก่อนสร้าง
@router.post("/users", status_code=201)
def create_user(user: UserCreate, request: Request, authorization: str = Header(None)):
    db, actor = _require_super_admin(authorization)
    email = user.email
    if db.collection("users").document(email).get().exists:
        raise HTTPException(status_code=409, detail="Email นี้มีอยู่ในระบบแล้ว")
    user_dict = user.model_dump()
    #? Hash รหัสผ่านก่อนบันทึก — ไม่เก็บ plaintext ลง Firestore เด็ดขาด
    user_dict["password"] = _hash_password(user_dict["password"])
    db.collection("users").document(email).set(user_dict)
    log_activity(db, action=LogAction.USER_CREATE, request=request, user=actor,
                 resource_id=email, resource_type="user",
                 details={"email": email, "level": user.level, "department": user.department})
    return {"success": True, "email": email}


@router.put("/users/{email:path}")
def update_user(email: str, update: UserUpdate, request: Request, authorization: str = Header(None)):
    db, actor = _require_super_admin(authorization)
    doc_ref = db.collection("users").document(email)
    if not doc_ref.get().exists:
        raise HTTPException(status_code=404, detail="User not found")
    update_data = update.model_dump(exclude_unset=True)
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    #? ถ้ามีการส่ง password มาด้วย ให้ hash ก่อนบันทึก
    if "password" in update_data and update_data["password"]:
        update_data["password"] = _hash_password(update_data["password"])
    doc_ref.update(update_data)
    #? แยก log: toggle ignore vs แก้ไขข้อมูลทั่วไป
    if list(update_data.keys()) == ["ignore"]:
        log_activity(db, action=LogAction.USER_TOGGLE, request=request, user=actor,
                     resource_id=email, resource_type="user",
                     details={"ignore": update_data["ignore"]})
    else:
        log_activity(db, action=LogAction.USER_UPDATE, request=request, user=actor,
                     resource_id=email, resource_type="user",
                     details={"updated_fields": [k for k in update_data if k != "password"]})
    return {"success": True}


@router.delete("/users/{email:path}")
def delete_user(email: str, request: Request, authorization: str = Header(None)):
    db, actor = _require_super_admin(authorization)
    doc_ref = db.collection("users").document(email)
    if not doc_ref.get().exists:
        raise HTTPException(status_code=404, detail="User not found")
    doc_ref.delete()
    log_activity(db, action=LogAction.USER_DELETE, request=request, user=actor,
                 resource_id=email, resource_type="user", details={"email": email})
    return {"success": True}


class EvaluationUpdate(BaseModel):
    evaluator_ids: List[str]


#? ดึงข้อมูลสายการบังคับบัญชา (ใครประเมินใคร)
@router.get("/evaluations")
def list_evaluations(authorization: str = Header(None)):
    """Get all users with their evaluators (join from evaluations collection)"""
    db, _ = _require_super_admin(authorization)

    users_by_pid = {}
    for doc in db.collection("users").stream():
        u = doc.to_dict()
        pid = u.get("personal_id")
        if pid:
            users_by_pid[pid] = {
                "personal_id": pid,
                "name": f"{u.get('firstname','')} {u.get('lastname','')}".strip(),
                "position": u.get("position", ""),
                "department": u.get("department", ""),
                "level": u.get("level", 0),
                "ignore": u.get("ignore", 0),
            }

    eval_map = {}
    for doc in db.collection("evaluations").stream():
        d = doc.to_dict()
        target_id = d.get("target_id", doc.id)
        evaluator_list = []
        for ev in sorted(d.get("evaluators", []), key=lambda x: x.get("order", 0)):
            ev_id = ev.get("evaluator_id")
            ev_info = users_by_pid.get(ev_id, {"name": ev_id, "position": "", "department": ""})
            evaluator_list.append({
                "evaluator_id": ev_id,
                "name": ev_info["name"],
                "position": ev_info["position"],
                "department": ev_info["department"],
            })
        eval_map[target_id] = evaluator_list

    all_users_list = list(users_by_pid.values())
    all_users_list.sort(key=lambda x: (x.get("department", ""), x.get("name", "")))

    evaluations = []
    for pid, u in users_by_pid.items():
        evaluations.append({
            "target_id": pid,
            "target_name": u["name"],
            "target_department": u["department"],
            "target_position": u["position"],
            "target_level": u["level"],
            "target_ignore": u["ignore"],
            "evaluators": eval_map.get(pid, []),
        })
    evaluations.sort(key=lambda x: (x.get("target_department", ""), x.get("target_name", "")))

    return {"users": all_users_list, "evaluations": evaluations}


@router.put("/evaluations/{target_id}")
def update_evaluation(target_id: str, update: EvaluationUpdate, request: Request, authorization: str = Header(None)):
    """Create/update evaluators for a target (replaces the full list)"""
    db, actor = _require_super_admin(authorization)
    doc_ref = db.collection("evaluations").document(target_id)
    evaluators = [{"evaluator_id": pid, "order": i + 1} for i, pid in enumerate(update.evaluator_ids)]
    doc_ref.set({
        "target_id": target_id,
        "evaluators": evaluators,
        "evaluator_ids": update.evaluator_ids,
    })
    log_activity(db, action=LogAction.EVALUATION_UPDATE, request=request, user=actor,
                 resource_id=target_id, resource_type="evaluation",
                 details={"evaluator_count": len(update.evaluator_ids), "evaluator_ids": update.evaluator_ids})
    return {"success": True}


@router.post("/migrate/ignore")
def migrate_ignore(authorization: str = Header(None)):
    """เพิ่ม ignore=0 ให้ผู้ใช้ที่ยังไม่มี field นี้"""
    db, _ = _require_super_admin(authorization)
    count = 0
    for doc in db.collection("users").stream():
        if 'ignore' not in doc.to_dict():
            doc.reference.update({'ignore': 0})
            count += 1
    return {"updated": count}


@router.post("/migrate/evaluator-ids")
def migrate_evaluator_ids(authorization: str = Header(None)):
    """เพิ่ม evaluator_ids (flat array) ให้ evaluation documents ที่ยังไม่มี field นี้"""
    db, _ = _require_super_admin(authorization)
    count = 0
    for doc in db.collection("evaluations").stream():
        d = doc.to_dict()
        if 'evaluator_ids' not in d:
            ids = [ev['evaluator_id'] for ev in d.get('evaluators', [])]
            doc.reference.update({'evaluator_ids': ids})
            count += 1
    return {"updated": count}


# ── Activity Logs ────────────────────────────────────────────────────────────

@router.get("/logs")
def get_activity_logs(
    request: Request,
    category: Optional[str] = None,
    user_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    authorization: str = Header(None),
):
    """ดึง activity logs — super admin เท่านั้น
    Filter: category, user_id, date_from (YYYY-MM-DD), date_to (YYYY-MM-DD)"""
    db, _ = _require_super_admin(authorization)
    if limit > 500:
        limit = 500

    query = db.collection("activity_logs").order_by("timestamp", direction="DESCENDING")

    #? ดึงมาก่อนแล้ว filter ใน Python เพราะ Firestore ไม่รองรับ range query บนหลาย field พร้อมกัน
    all_docs = list(query.stream())

    results = []
    for doc in all_docs:
        d = doc.to_dict()
        ts = d.get("timestamp", "")

        if category and d.get("category") != category:
            continue
        if user_id and d.get("user_id") != user_id:
            continue
        if date_from and ts[:10] < date_from:
            continue
        if date_to and ts[:10] > date_to:
            continue

        results.append({"id": doc.id, **d})

    total = len(results)
    paginated = results[offset: offset + limit]
    return {"total": total, "logs": paginated}


# ── Monthly Stats ────────────────────────────────────────────────────────────

#? ฟังก์ชันคำนวณสถิติรายเดือน (ใช้สำหรับการแสดงผลบนเว็บและ Export Excel)
def _compute_monthly_stats(db, month: str, allowed_user_ids=None) -> dict:
    """Compute per-user monthly statistics. month = 'YYYY-MM'.
    If allowed_user_ids is a set, only include those users.
    If None, include all active users (super admin)."""
    if not re.match(r'^\d{4}-\d{2}$', month):
        raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")

    year, mon = int(month[:4]), int(month[5:7])
    _, num_days = calendar.monthrange(year, mon)
    #? คำนวณจำนวนวันทำงาน (จันทร์-ศุกร์) ของเดือนนั้นๆ
    weekdays = sum(1 for d in range(1, num_days + 1) if date_mod(year, mon, d).weekday() < 5)

    # Fetch active users (filtered by allowed_user_ids if provided)
    users_map = {}
    for doc in db.collection("users").stream():
        u = doc.to_dict()
        if u.get("ignore", 0) == 0:
            pid = u.get("personal_id")
            if pid:
                if allowed_user_ids is not None and pid not in allowed_user_ids:
                    continue
                users_map[pid] = {
                    "personal_id": pid,
                    "name": f"{u.get('firstname', '')} {u.get('lastname', '')}".strip(),
                    "position": u.get("position", ""),
                    "department": u.get("department", ""),
                }

    # Accumulate per-user stats from reports
    accum = {
        pid: {"days_submitted": 0, "total_progress": 0,
              "wfh_days": 0, "onsite_days": 0, "hybrid_days": 0,
              "total_tasks": 0, "done_tasks": 0, "problem_days": 0}
        for pid in users_map
    }

    #? กรองเฉพาะ reports ของเดือนที่ต้องการที่ Firestore — ไม่ดึง reports ทั้งหมดตลอดกาล
    next_mon = mon + 1 if mon < 12 else 1
    next_yr  = year if mon < 12 else year + 1
    reports_query = (
        db.collection("reports")
          .where("timestamp", ">=", f"{month}-01")
          .where("timestamp", "<",  f"{next_yr:04d}-{next_mon:02d}-01")
    )
    for doc in reports_query.stream():
        r = doc.to_dict()
        uid = r.get("user_id")
        if uid not in accum:
            continue
        s = accum[uid]
        s["days_submitted"] += 1
        s["total_progress"] += r.get("progress", 0)
        wm = r.get("work_mode", "").lower()
        if wm == "wfh":
            s["wfh_days"] += 1
        elif wm == "onsite":
            s["onsite_days"] += 1
        elif wm == "hybrid":
            s["hybrid_days"] += 1
        tasks = r.get("tasks", [])
        s["total_tasks"] += len(tasks)
        s["done_tasks"] += sum(1 for t in tasks if t.get("status") == "done")
        prob = r.get("problems", "-") or "-"
        if prob.strip() not in ("-", ""):
            s["problem_days"] += 1

    users_list = []
    for pid, uinfo in users_map.items():
        s = accum[pid]
        days = s["days_submitted"]
        avg_prog = round(s["total_progress"] / days, 1) if days > 0 else 0
        compliance = round(days / weekdays * 100, 1) if weekdays > 0 else 0
        users_list.append({
            **uinfo,
            "days_submitted": days,
            "compliance": compliance,
            "avg_progress": avg_prog,
            "wfh_days": s["wfh_days"],
            "onsite_days": s["onsite_days"],
            "hybrid_days": s["hybrid_days"],
            "total_tasks": s["total_tasks"],
            "done_tasks": s["done_tasks"],
            "problem_days": s["problem_days"],
        })

    users_list.sort(key=lambda x: (x["department"], x["name"]))
    return {"month": month, "calendar_days": num_days, "weekdays": weekdays, "users": users_list}


@router.get("/stats")
def get_stats(month: str, request: Request, authorization: str = Header(None)):
    """สถิติรายเดือน — ส่งคืน per-user stats จัดกลุ่มตาม department
    Super admin เห็นทุกคน, supervisor เห็นเฉพาะลูกน้องในสายบังคับบัญชา"""
    db, user_data = _require_any_admin(authorization)
    allowed = _get_visible_user_ids(db, user_data)
    result = _compute_monthly_stats(db, month, allowed)
    log_activity(db, action=LogAction.STATS_VIEW, request=request, user=user_data,
                 details={"month": month})
    return result


#? Export สถิติรายงานตัวประจำเดือนเป็นไฟล์ Excel (.xlsx)
@router.get("/stats/export")
def export_stats(month: str, request: Request, authorization: str = Header(None)):
    """Export สถิติรายเดือนเป็นไฟล์ Excel (.xlsx)
    Super admin ได้ทุกคน, supervisor ได้เฉพาะลูกน้อง"""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    db, user_data = _require_any_admin(authorization)
    allowed = _get_visible_user_ids(db, user_data)
    data = _compute_monthly_stats(db, month, allowed)
    log_activity(db, action=LogAction.STATS_EXPORT, request=request, user=user_data,
                 details={"month": month})

    # Thai month names
    th_months = ["", "ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
                 "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
    year_val, mon_val = int(month[:4]), int(month[5:7])
    month_label = f"{th_months[mon_val]} {year_val + 543}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"สถิติ {month}"

    # ── Styles ──
    def side():
        return Side(style="thin", color="CCCCCC")

    def cell_border():
        return Border(left=side(), right=side(), top=side(), bottom=side())

    hdr_fill = PatternFill("solid", fgColor="1059A3")
    dept_fill = PatternFill("solid", fgColor="E8EFF8")
    green_fill = PatternFill("solid", fgColor="D4EDDA")
    yellow_fill = PatternFill("solid", fgColor="FFF3CD")
    red_fill = PatternFill("solid", fgColor="F8D7DA")
    alt_fill = PatternFill("solid", fgColor="F9FAFB")

    #todo ย้าย Style การจัดไฟล์ Excel ไปไว้ใน Utility แยกต่างหากเพื่อลดความยาวโค้ด
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    dept_font = Font(bold=True, color="1A3A6B", size=10)
    body_font = Font(size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Title ──
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = f"สถิติการส่งรายงานประจำเดือน {month_label}"
    title_cell.font = Font(bold=True, size=13, color="1059A3")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:N2")
    sub_cell = ws["A2"]
    sub_cell.value = (f"วันทำงาน (จันทร์–ศุกร์): {data['weekdays']} วัน  |  "
                      f"ผู้ใช้งานทั้งหมด: {len(data['users'])} คน  |  "
                      f"สร้างวันที่: {date_mod.today().strftime('%d/%m/')}{date_mod.today().year+543}")
    sub_cell.font = Font(size=9, color="64748B")
    sub_cell.alignment = left
    ws.row_dimensions[2].height = 18

    # ── Header row ──
    headers = [
        ("ลำดับ", 6), ("ชื่อ-สกุล", 24), ("ตำแหน่ง", 20), ("หน่วยงาน", 18),
        ("ส่งรายงาน\n(วัน)", 10), ("วันทำงาน\n(วัน)", 10), ("Compliance\n(%)", 11),
        ("Avg Progress\n(%)", 12), ("WFH\n(วัน)", 8), ("On-site\n(วัน)", 8),
        ("Hybrid\n(วัน)", 8), ("งาน\n(ทั้งหมด)", 9), ("งาน\n(เสร็จ)", 9),
        ("มีปัญหา\n(วัน)", 9),
    ]
    col_letters = [chr(65 + i) for i in range(len(headers))]
    for i, (title, width) in enumerate(headers):
        c = ws.cell(row=3, column=i + 1, value=title)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = cell_border()
        ws.column_dimensions[col_letters[i]].width = width
    ws.row_dimensions[3].height = 32

    # ── Data rows ──
    row_num = 4
    seq = 0

    # Group by department
    depts = {}
    for u in data["users"]:
        dept = u["department"] or "ไม่ระบุหน่วยงาน"
        depts.setdefault(dept, []).append(u)

    for dept, members in depts.items():
        # Department header row
        ws.merge_cells(f"A{row_num}:N{row_num}")
        dc = ws.cell(row=row_num, column=1, value=f"  {dept}  ({len(members)} คน)")
        dc.font = dept_font
        dc.fill = dept_fill
        dc.alignment = left
        dc.border = cell_border()
        ws.row_dimensions[row_num].height = 18
        row_num += 1

        for idx, u in enumerate(members):
            seq += 1
            fill = alt_fill if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            comp = u["compliance"]
            prog = u["avg_progress"]
            comp_fill = green_fill if comp >= 80 else (yellow_fill if comp >= 50 else red_fill)
            prog_fill = green_fill if prog >= 80 else (yellow_fill if prog >= 40 else red_fill)

            row_data = [
                seq, u["name"], u["position"], u["department"],
                u["days_submitted"], data["weekdays"],
                comp, prog,
                u["wfh_days"], u["onsite_days"], u["hybrid_days"],
                u["total_tasks"], u["done_tasks"], u["problem_days"],
            ]
            for col_idx, val in enumerate(row_data):
                c = ws.cell(row=row_num, column=col_idx + 1, value=val)
                c.font = body_font
                c.border = cell_border()
                c.alignment = left if col_idx in (1, 2, 3) else center
                if col_idx == 6:  # compliance
                    c.fill = comp_fill
                elif col_idx == 7:  # avg progress
                    c.fill = prog_fill
                else:
                    c.fill = fill

            ws.row_dimensions[row_num].height = 16
            row_num += 1

    # Freeze header rows
    ws.freeze_panes = "A4"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"stats_{month}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Daily Report Export ──────────────────────────────────────────────────────

#? Export รายละเอียดรายงานรายวันของทุกคน (หรือลูกน้อง) เป็นไฟล์ Excel
@router.get("/reports/export")
def export_daily_reports(date: str, request: Request, authorization: str = Header(None)):
    """Export รายงานประจำวันเป็นไฟล์ Excel (.xlsx) — admin level 1+, กรองตามสิทธิ์
    date = 'YYYY-MM-DD'"""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not re.match(r'^\d{4}-\d{2}-\d{2}$', date):
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    db, user_data = _require_any_admin(authorization)
    allowed = _get_visible_user_ids(db, user_data)
    log_activity(db, action=LogAction.REPORT_EXPORT, request=request, user=user_data,
                 details={"date": date})

    # Fetch all active users
    users_map = {}
    for doc in db.collection("users").stream():
        u = doc.to_dict()
        if u.get("ignore", 0) == 0:
            pid = u.get("personal_id")
            if pid:
                if allowed is not None and pid not in allowed:
                    continue
                users_map[pid] = {
                    "personal_id": pid,
                    "name": f"{u.get('firstname', '')} {u.get('lastname', '')}".strip(),
                    "position": u.get("position", ""),
                    "department": u.get("department", ""),
                }

    # Fetch reports for the given date
    reports_map = {}
    for doc in db.collection("reports").stream():
        r = doc.to_dict()
        if not r.get("timestamp", "").startswith(date):
            continue
        uid = r.get("user_id")
        if uid in users_map:
            reports_map[uid] = r

    # ── Thai formatting ──
    th_months = ["", "ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
                 "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
    th_days = ["จันทร์", "อังคาร", "พุธ", "พฤหัสบดี", "ศุกร์", "เสาร์", "อาทิตย์"]
    d = date_mod(int(date[:4]), int(date[5:7]), int(date[8:10]))
    thai_date = f"วัน{th_days[d.weekday()]}ที่ {d.day} {th_months[d.month]} {d.year + 543}"

    status_map = {"done": "✓ เสร็จสิ้น", "prog": "⋯ กำลังดำเนินการ", "pend": "◯ รอดำเนินการ"}
    mode_map = {"wfh": "WFH (ทำงานที่บ้าน)", "onsite": "On-site (ทำงานที่สำนักงาน)", "hybrid": "Hybrid (ทำทั้งที่บ้านและที่ทำงาน)"}

    # ── Build Excel ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"รายงาน {date}"

    def side_thin():
        return Side(style="thin", color="CCCCCC")

    def cell_border():
        return Border(left=side_thin(), right=side_thin(), top=side_thin(), bottom=side_thin())

    hdr_fill = PatternFill("solid", fgColor="1059A3")
    dept_fill = PatternFill("solid", fgColor="E8EFF8")
    unsent_fill = PatternFill("solid", fgColor="FCEBEB")
    green_fill = PatternFill("solid", fgColor="D4EDDA")
    yellow_fill = PatternFill("solid", fgColor="FFF3CD")
    red_fill = PatternFill("solid", fgColor="F8D7DA")
    alt_fill = PatternFill("solid", fgColor="F9FAFB")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    dept_font = Font(bold=True, color="1A3A6B", size=10)
    body_font = Font(size=10)
    body_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_center = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Title ──
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"สรุปรายงานประจำวัน {thai_date}"
    title_cell.font = Font(bold=True, size=13, color="1059A3")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    sent_count = len(reports_map)
    unsent_count = len(users_map) - sent_count
    ws.merge_cells("A2:L2")
    sub_cell = ws["A2"]
    sub_cell.value = (f"ส่งแล้ว: {sent_count} คน  |  ยังไม่ส่ง: {unsent_count} คน  |  "
                      f"ทั้งหมด: {len(users_map)} คน  |  "
                      f"สร้างวันที่: {date_mod.today().strftime('%d/%m/')}{date_mod.today().year+543}")
    sub_cell.font = Font(size=9, color="64748B")
    sub_cell.alignment = left_center
    ws.row_dimensions[2].height = 18

    # ── Header row ──
    headers = [
        ("ลำดับ", 6), ("ชื่อ-สกุล", 24), ("ตำแหน่ง", 18), ("หน่วยงาน", 18),
        ("รูปแบบ\nทำงาน", 14), ("ความคืบหน้า\n(%)", 11),
        ("รายการงาน", 40), ("สถานะงาน", 15),
        ("ปัญหา / อุปสรรค", 28), ("แผนวันพรุ่งนี้", 28),
        ("คอมเมนต์", 30), ("เวลาส่ง", 10),
    ]
    col_letters = [chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26) for i in range(len(headers))]
    for i, (title, width) in enumerate(headers):
        c = ws.cell(row=3, column=i + 1, value=title)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = cell_border()
        ws.column_dimensions[col_letters[i]].width = width
    ws.row_dimensions[3].height = 32

    # ── Data rows (sent) ──
    row_num = 4
    seq = 0

    # Group sent users by department
    sent_by_dept = {}
    for pid, uinfo in users_map.items():
        if pid in reports_map:
            dept = uinfo["department"] or "ไม่ระบุหน่วยงาน"
            sent_by_dept.setdefault(dept, []).append((uinfo, reports_map[pid]))

    for dept in sorted(sent_by_dept.keys()):
        members = sent_by_dept[dept]
        # Dept header
        ws.merge_cells(f"A{row_num}:L{row_num}")
        dc = ws.cell(row=row_num, column=1, value=f"  📁 {dept}  ({len(members)} คน ส่งแล้ว)")
        dc.font = dept_font
        dc.fill = dept_fill
        dc.alignment = left_center
        dc.border = cell_border()
        ws.row_dimensions[row_num].height = 20
        row_num += 1

        for idx, (uinfo, report) in enumerate(members):
            seq += 1
            fill = alt_fill if idx % 2 == 0 else white_fill

            tasks = report.get("tasks", [])
            
            def format_task(i, t):
                base = f"{i+1}. {t.get('title', '-')}"
                f_cnt = len(t.get('files', []))
                l_cnt = len(t.get('links', []))
                if f_cnt > 0 or l_cnt > 0:
                    base += f" [📎 {f_cnt} ไฟล์, 🔗 {l_cnt} ลิงก์]"
                return base

            task_lines = "\n".join([format_task(i, t) for i, t in enumerate(tasks)]) or "-"
            status_lines = "\n".join(
                [status_map.get(t.get("status", ""), t.get("status", "-")) for t in tasks]
            ) or "-"

            problems = report.get("problems", "-") or "-"
            plan = report.get("plan_tomorrow", "-") or "-"
            progress = report.get("progress", 0)
            work_mode = mode_map.get(report.get("work_mode", ""), report.get("work_mode", "-"))
            submit_time = report.get("submit_time", "-")

            # Comments
            comments = report.get("comments", [])
            comment_lines = "\n".join(
                [f"[{c.get('author_name', '-')}] {c.get('message', '')}"
                 + (f" #{c['tag']}" if c.get('tag') else "")
                 for c in comments]
            ) or "-"

            prog_fill = green_fill if progress >= 80 else (yellow_fill if progress >= 50 else red_fill)

            row_data = [
                seq, uinfo["name"], uinfo["position"], uinfo["department"],
                work_mode, progress,
                task_lines, status_lines,
                problems, plan, comment_lines, submit_time,
            ]
            max_lines = max(
                task_lines.count("\n") + 1,
                status_lines.count("\n") + 1,
                comment_lines.count("\n") + 1,
                1
            )
            for col_idx, val in enumerate(row_data):
                c = ws.cell(row=row_num, column=col_idx + 1, value=val)
                c.font = body_font
                c.border = cell_border()
                c.alignment = body_wrap
                if col_idx == 5:  # progress
                    c.fill = prog_fill
                    c.alignment = center
                elif col_idx == 0:  # seq
                    c.alignment = center
                    c.fill = fill
                else:
                    c.fill = fill

            ws.row_dimensions[row_num].height = max(16, min(max_lines * 15, 120))
            row_num += 1

    # ── Unsent section ──
    unsent_users = [u for pid, u in users_map.items() if pid not in reports_map]
    if unsent_users:
        row_num += 1
        ws.merge_cells(f"A{row_num}:L{row_num}")
        unsent_hdr = ws.cell(row=row_num, column=1,
                             value=f"  ❌ ยังไม่ส่งรายงาน ({len(unsent_users)} คน)")
        unsent_hdr.font = Font(bold=True, color="791F1F", size=11)
        unsent_hdr.fill = unsent_fill
        unsent_hdr.alignment = left_center
        unsent_hdr.border = cell_border()
        ws.row_dimensions[row_num].height = 22
        row_num += 1

        # Group unsent by department
        unsent_by_dept = {}
        for u in unsent_users:
            dept = u["department"] or "ไม่ระบุหน่วยงาน"
            unsent_by_dept.setdefault(dept, []).append(u)

        for dept in sorted(unsent_by_dept.keys()):
            members = unsent_by_dept[dept]
            ws.merge_cells(f"A{row_num}:L{row_num}")
            dc = ws.cell(row=row_num, column=1, value=f"  📁 {dept}  ({len(members)} คน)")
            dc.font = dept_font
            dc.fill = dept_fill
            dc.alignment = left_center
            dc.border = cell_border()
            ws.row_dimensions[row_num].height = 18
            row_num += 1

            for idx, u in enumerate(members):
                seq += 1
                fill = PatternFill("solid", fgColor="FFF5F5") if idx % 2 == 0 else unsent_fill
                row_data = [
                    seq, u["name"], u["position"], u["department"],
                    "-", "-", "-", "-", "-", "-", "-", "ยังไม่ส่ง",
                ]
                for col_idx, val in enumerate(row_data):
                    c = ws.cell(row=row_num, column=col_idx + 1, value=val)
                    c.font = Font(size=10, color="791F1F") if col_idx == 11 else body_font
                    c.border = cell_border()
                    c.alignment = center if col_idx in (0, 5, 11) else body_wrap
                    c.fill = fill
                ws.row_dimensions[row_num].height = 16
                row_num += 1

    # Freeze header
    ws.freeze_panes = "A4"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"daily_report_{date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
